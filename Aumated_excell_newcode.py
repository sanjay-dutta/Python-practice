# The aim of the program is to aumation on the excel sheet. #
# Here we need to multiply the rows of colum 2 values with 0.9 in a excel sheet

# import openpyxl package
import openpyxl as xl
# import anther package for barchart
from openpyxl.chart import BarChart, Reference 

def process_workbook(filename):
    # Loading excel workbook
    wb = xl.load_workbook(filename)
    # in the workbook we have only one sheet
    sheet = wb["Sheet1"]
    
    for row in range(2, sheet.max_row + 1):
        # Access to the column 3
        cell = sheet.cell(row, 3)
        # multiply the value of the column 3 values with 0.9
        corrected_price = cell.value * 0.9
        # Add new cell the worksheet
        corrected_price_cell = sheet.cell(row, 4)
        # Set the value the corrected price
        corrected_price_cell.value = corrected_price

    # Select a reference class to select a range of values
    values  = Reference(sheet,
        min_row = 2, 
        max_row = sheet.max_row,
        min_col = 4,
        max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    # Create and save a new excel sheet
    wb.save(filename)